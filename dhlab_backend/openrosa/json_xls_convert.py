import re

from django.http import HttpResponse
from openpyxl import Workbook

from openpyxl.writer.excel import save_virtual_workbook

class jsonXlsConvert():
	def __init__(self, file_name): 
		jsonXlsConvert.wb = Workbook()
		jsonXlsConvert.worksheet1 = self.wb.get_active_sheet()
		self.worksheet1.title = 'survey'
		jsonXlsConvert.surveyHeaders = ['type', 'name', 'label', 'hint', 'media::image',
						 'media::video', 'media::audio', 'constraint', 
						 'constraint_message', 'required', 'default', 
						 'relevant', 'read_only', 'calculation', 'appearance']
		self.writeArrayToXLS(self.worksheet1, self.surveyHeaders, 'A1')

		jsonXlsConvert.worksheet2 = self.wb.create_sheet(1)
		self.worksheet2.title = 'choices'
		jsonXlsConvert.choicesHeaders = ['list name', 'name', 'label', 'image']
		self.writeArrayToXLS(self.worksheet2, self.choicesHeaders, 'A1')
		jsonXlsConvert.file_name = (file_name + ".xlsx")

		jsonXlsConvert.iterRow = 1
		jsonXlsConvert.choicesRow = 1

	def writeToXls(self, json_dict):

		self.recurseWriter(json_dict)

		response = HttpResponse(save_virtual_workbook(self.wb), content_type='application/vnd.ms-excel')
		response['Content-Disposition'] = 'attachment; filename="' + self.file_name + '"'
		print self.choicesHeaders
		return response

	def recurseWriter(self, json_dict):
		for element in json_dict:
			self.iterRow += 1
			surveyWriteRow = []

			if element.get("type") == 'group':
				self.iterRow += 1
				surveyWriteRow = ['begin group']
				self.writeArrayToXLS(self.worksheet1, surveyWriteRow, ('A' + `self.iterRow`))
				self.recurseWriter(element.get("children"))
				self.iterRow += 1
				surveyWriteRow = ['end group']
				self.writeArrayToXLS(self.worksheet1, surveyWriteRow, ('A' + `self.iterRow`))
				self.iterRow += 1

			else:
				for header in self.surveyHeaders:
					if header in element:
						surveyWriteRow += [element.get(header)]
					else:
						surveyWriteRow += [""]
				self.writeArrayToXLS(self.worksheet1, surveyWriteRow, ('A' + `self.iterRow`))

			if 'choices' in element:
				choiceOptions = element['choices']
				self.choicesRow += 1
				for choice in choiceOptions:
					choiceWriteRow = [element.get("name")]
					for header in self.choicesHeaders:
						if header in choice:
							choiceWriteRow += [choice.get(header)]
						else:
							choiceWriteRow += [""] 
					self.writeArrayToXLS(self.worksheet2, choiceWriteRow, ('A' + `self.choicesRow`))
					self.choicesRow += 1

	'''Utility function to write an array to excel using openpyxl'''
	def writeArrayToXLS(self, worksheet, array_to_write, starting_cell, horizontal=True):
		res = re.findall(r'\d+', starting_cell)

		row = res[0]
		col = starting_cell[:-len(res[0])]
		col_ind = ord(col) - 96 + 32 # only works for A-Z in caps

		if horizontal:
			for i in range(0,len(array_to_write)):
				cell_str = '%s%d' % (chr(col_ind + 64 + i), int(row))
				worksheet.cell(cell_str).value = array_to_write[i]
				# print cell_str

		else:
			for i in range(0,len(array_to_write)):
				cell_str = '%s%d' % (chr(col_ind + 64 ), int(row) + i)
				worksheet.cell(cell_str).value = array_to_write[i]
				# print cell_str
