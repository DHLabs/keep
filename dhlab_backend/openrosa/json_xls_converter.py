# This source contains snippets of xml2json 

"""xml2json.py  Convert XML to JSON

Relies on ElementTree for the XML parsing.  This is based on
pesterfish.py but uses a different XML->JSON mapping.
The XML->JSON mapping is described at
http://www.xml.com/pub/a/2006/05/31/converting-between-xml-and-json.html

Rewritten to a command line utility by Hay Kranen < github.com/hay >

XML                              JSON
<e/>                             "e": null
<e>text</e>                      "e": "text"
<e name="value" />               "e": { "@name": "value" }
<e name="value">text</e>         "e": { "@name": "value", "#text": "text" }
<e> <a>text</a ><b>text</b> </e> "e": { "a": "text", "b": "text" }
<e> <a>text</a> <a>text</a> </e> "e": { "a": ["text", "text"] }
<e> text <a>text</a> </e>        "e": { "#text": "text", "a": "text" }

This is very similar to the mapping used for Yahoo Web Services
(http://developer.yahoo.com/common/json.html#xml).

This is a mess in that it is so unpredictable -- it requires lots of testing
(e.g. to see if values are lists or strings or dictionaries).  For use
in Python this could be vastly cleaner.  Think about whether the internal
form can be more self-consistent while maintaining good external characteristics
for the JSON.

Look at the Yahoo version closely to see how it works.  Maybe can adopt
that completely if it makes more sense...

R. White, 2006 November 6
"""

import xml.etree.cElementTree as ET
from xml.etree.cElementTree import parse
import simplejson, optparse, sys, os

def elem_to_internal(elem,strip=1):

    """Convert an Element into an internal dictionary (not JSON!)."""

    d = {}
    for key, value in elem.attrib.items():
        d['@'+key] = value

    # loop over subelements to merge them
    for subelem in elem:
        v = elem_to_internal(subelem,strip=strip)
        tag = subelem.tag
        value = v[tag]
        try:
            # add to existing list for this tag
            d[tag].append(value)
        except AttributeError:
            # turn existing entry into a list
            d[tag] = [d[tag], value]
        except KeyError:
            # add a new non-list entry
            d[tag] = value
    text = elem.text
    tail = elem.tail
    if strip:
        # ignore leading and trailing whitespace
        if text: text = text.strip()
        if tail: tail = tail.strip()

    if tail:
        d['#tail'] = tail

    if d:
        # use #text element if other attributes exist
        if text: d["#text"] = text
    else:
        # text is the value if no attributes
        d = text or None
    return {elem.tag: d}



def elem2json(elem, strip=1):

    """Convert an ElementTree or Element into a JSON string."""

    if hasattr(elem, 'getroot'):
        elem = elem.getroot()
    return simplejson.dumps(elem_to_internal(elem,strip=strip), sort_keys=False)



def xml2json(xml_file,strip=1):

    """Convert an XML string into a JSON string."""

    # elem = ET
    xmlfile = open(xml_file, "r")
    tree = parse(xmlfile)
    elem = tree.getroot()

    return elem2json(elem,strip=strip)




import json, pprint
import openpyxl, xmltodict
from openpyxl import Workbook




'''Utility function to write na array to excel using openpyxl'''
def writeArrayToXLS(worksheet, array_to_write, starting_cell, horizontal=True):

    import re
    res = re.findall(r'\d+', starting_cell)
    
    row = res[0]
    col = starting_cell[:-len(res[0])]
    col_ind = ord(col) - 96 + 32 # only works for A-Z in caps

    if horizontal:
        for i in range(0,len(array_to_write)):
            cell_str = '%s%d' % (chr(col_ind + 64 + i), int(row))
            worksheet.cell(cell_str).value = array_to_write[i]
            # print cell_str

    else:
        for i in range(0,len(array_to_write)):
            cell_str = '%s%d' % (chr(col_ind + 64 ), int(row) + i)
            worksheet.cell(cell_str).value = array_to_write[i]
            # print cell_str


'''main class'''
class xformConvert(object):
    """docstring for xformConvert"""
    def __init__(self, file_name):
        super(xformConvert, self).__init__()
        # self.arg = arg

        self.XF_model    = "{http://www.w3.org/2002/xforms}model"
        self.XF_bind     = "{http://www.w3.org/2002/xforms}bind"
        self.XF_itext    = "{http://www.w3.org/2002/xforms}itext"
        self.XF_translation  = "{http://www.w3.org/2002/xforms}translation"
        self.XF_text     = "{http://www.w3.org/2002/xforms}text"
        self.XF_instance = "{http://www.w3.org/2002/xforms}instance"
        self.XF_value    = "{http://www.w3.org/2002/xforms}value"

        self.JR_constraint   = "@{http://openrosa.org/javarosa}constraintMsg"

        self.XF_head     = "{http://www.w3.org/1999/xhtml}head"
        self.XF_body     = "{http://www.w3.org/1999/xhtml}body"
        self.XF_html     = "{http://www.w3.org/1999/xhtml}html"
        self.XF_title    = "{http://www.w3.org/1999/xhtml}title"

        self.XF_hint     = '{http://www.w3.org/2002/xforms}hint'
        self.XF_label     = '{http://www.w3.org/2002/xforms}label'
        self.XF_item     = '{http://www.w3.org/2002/xforms}item'    

        
        self.output_file = file_name.split('.')[0] + '.xlsx'
        self.jsondict = json.loads(xml2json(file_name, strip=0))

        self.group_labels = []

        self.header_dict = self.jsondict[self.XF_html][self.XF_head]
        self.bind_dict = self.header_dict[self.XF_model][self.XF_bind]
        
        self.hasTranslations = True
        try:
            self.itext_dict = self.header_dict[self.XF_model][self.XF_itext]
        except:
            self.hasTranslations = False
            print 'no itext tag found, assuming zero translations.'

        self.instance_dict = self.header_dict[self.XF_model][self.XF_instance]

        self.body_dict = self.jsondict[self.XF_html][self.XF_body]
        self.title = self.header_dict[self.XF_title]['#text']

        # for chcoices
        self.currentRow = 2

        # print 'loaded Form with name: %s' % self.title

    ###########################################################
    # Translation
    ###########################################################
    def parseTranslations(self):
        
        self.number_of_translations = 0

        if self.hasTranslations:

            translation_obj = self.itext_dict[self.XF_translation]
            translation_labels = {}

            if type(translation_obj) == dict:
                    translation_obj = [translation_obj]

            print 'there are %d translations' % (len(translation_obj))
            if len(translation_obj) > 0:

                print 'parsing translations'
                self.number_of_translations = len(translation_obj)

                translation_languages = []
                default_lang = ''

                for lang in translation_obj:
                    language_label = lang["@lang"]
                    translation_languages.append(language_label)
                    if "@default" in lang.keys():
                        if lang["@default"] == "true()":
                            default_lang = language_label

                    trans_list = lang[self.XF_text]
                    # print text_list
                    if type(trans_list) == dict:
                        trans_list = [trans_list]

                    for trans_entry in trans_list:

                        if len(trans_entry) > 2:

                            trans_id = trans_entry["@id"]
                            if trans_id not in translation_labels.keys():

                                # doesn't exist, creating
                                translation_labels[trans_id] = {}
                                # print trans_entry[self.XF_value]
                                if type(trans_entry[self.XF_value]) == dict:

                                    if '#text' in trans_entry[self.XF_value].keys():
                                        translation_labels[trans_id][language_label] = trans_entry[self.XF_value]['#text']
                                        # print '\t Found, adding %s' % trans_entry[self.XF_value]['#text']
                                    else:
                                        # print trans_entry[self.XF_value].keys()
                                        translation_labels[trans_id][language_label] = ''
                                else:
                                    print 'not a dict in translation entry, indicates a special case (either image or similar)'
                                    translation_labels[trans_id][language_label] = '##UNDEF##'
                                    # print trans_entry[self.XF_value]
                            else:

                                if '#text' in trans_entry[self.XF_value].keys():
                                    translation_labels[trans_id][language_label] = trans_entry[self.XF_value]['#text']
                                    # print '\t Found, adding %s' % trans_entry[self.XF_value]['#text']
                                else:
                                    translation_labels[trans_id][language_label] = ''

                if default_lang == '':
                    # no default specified, use first
                    default_lang = translation_languages[0]

                self.translation_labels = translation_labels
                self.translation_languages = translation_languages
                self.default_lang = default_lang
                print 'default language: ' + self.default_lang    
            else:

                'no translations found, using embedded labels'
        else:
            'no translations found, using embedded labels'
            self.default_lang = ''



    ###########################################################
    # Build bind order
    ###########################################################
    def parseBindOrder(self):
        question_list = []
        element_count = len(self.bind_dict)

        for element in self.bind_dict:

            # print element
            question_dict = {}
            
            for key in element.keys():
                
                if key == "@nodeset":
                    question_dict['path'] = element[key]

                    path_array = element[key].split('/')
                    question_dict['path_root'] = path_array[1]
        
                    question_dict['element_name'] = path_array[-1]
                    if len(path_array) > 3:
                        question_dict['groups'] = path_array[2:-1]

                if key == "@calculate":
                    question_dict['calculate'] = element[key]

                if key == "@constraint":
                    question_dict['constraint'] = element[key]                

                if key == "@type":
                    question_dict['type'] = element[key]

                if key == "@readonly":
                    question_dict['readonly'] = element[key]

                if key == "@required":
                    question_dict['required'] = element[key]

                if key == self.JR_constraint:
                    question_dict['constraint_message'] = element[key]

            if 'type' not in question_dict.keys():
                question_dict['type'] = 'undefined'

            question_list.append(question_dict)

        self.question_list = question_list


    ###########################################################
    # Build get existing values from instance dict
    ###########################################################
    def parseInstanceData(self):

        # get base model
        form_keys = self.instance_dict.keys()
        assert len(form_keys) == 3
        for key in form_keys:
            if key not in ['#tail', '#text']:
                form_name = key

        instance_form_dict = self.instance_dict[form_name]

        # look guys, a real use for recursion!!
        def getNextDict(dict_to_use, key_array):
            if len(key_array) == 1:
                key_to_use = '{http://www.w3.org/2002/xforms}%s' % key_array[0]
                return dict_to_use[key_to_use]
            else:
                key_to_use = '{http://www.w3.org/2002/xforms}%s' % key_array[0]
                return getNextDict(dict_to_use[key_to_use], key_array[1:])

        # iterate through all questions and pre populate with data
        for question in self.question_list:
            if 'groups' in question.keys():
                instance_val = getNextDict(instance_form_dict, question['groups'] )
            else:
                key_to_use = '{http://www.w3.org/2002/xforms}%s' % question['element_name']
                instance_val = instance_form_dict[key_to_use]

            default_val = ''
            if '#text' in instance_val.keys():
                if len(instance_val['#text'].strip()) > 0:
                    default_val = instance_val['#text'].strip()

            question['default_val'] = default_val



    def parseBody(self):

        self.parseBodyRecurse(self.body_dict)

    def parseBodyRecurse(self, base_dict):

        XF_select1  = "{http://www.w3.org/2002/xforms}select1"
        XF_select   = "{http://www.w3.org/2002/xforms}select"
        XF_input    = "{http://www.w3.org/2002/xforms}input"
        XF_trigger  = "{http://www.w3.org/2002/xforms}trigger"
        XF_group    = "{http://www.w3.org/2002/xforms}group"
        XF_upload   = "{http://www.w3.org/2002/xforms}upload"

        list_to_search = [XF_select1, XF_select, XF_input, XF_trigger, XF_group, XF_upload]
        # list_to_search = [XF_group]

        for body_tag in list_to_search:

            if body_tag in base_dict.keys():

                tag_list = base_dict[body_tag]
                if type(tag_list) != list:
                    tag_list = [tag_list]

                if body_tag == XF_group:
                    for item in tag_list:
                        group_data = {}                    
                        for key in item:
                            if key == self.XF_label:

                                if type(item[key]) == dict:
                                    if '#text' in item[key].keys():
                                        group_data['text_label'] = item[key]['#text']
                                    if '@ref' in item[key].keys():
                                        group_data['trans_label'] = item[key]['@ref']

                                else:
                                    group_data['group-label'] = item[key]

                            if key == '@appearance':
                                group_data['group-layout'] = item[key]

                            if key == '@ref':
                                group_data['group-path'] = item[key]

                        self.group_labels.append(group_data)                                

                        self.parseBodyRecurse(item)

                else:

                    for item in tag_list:
                        # print item
                        for question in self.question_list:
                            if question['path'] == item['@ref']:
                                question['body_tag'] = body_tag

                                for key in item:
                                    if key == self.XF_hint:
                                        hint_dict = item[key]
                                        if '#text' in hint_dict.keys():
                                            question['text_hint'] = hint_dict['#text'].strip()
                                        if '@ref' in hint_dict.keys():
                                            question['trans_hint'] = hint_dict['@ref']

                                    if key == self.XF_label:
                                        label_dict = item[key]
                                        if '#text' in label_dict.keys():
                                            question['text_label'] = label_dict['#text'].strip()
                                        if '@ref' in label_dict.keys():
                                            question['trans_label'] = label_dict['@ref']

                                    if key == '@appearance':
                                        question['appearance'] = item[key]

                                    if key == self.XF_item:
                                        clean_list = []
                                        options_list = item[key]

                                        if type(options_list) != list:
                                            options_list = [options_list]

                                        for each_option in options_list:
                                            new_opt = {}

                                            for option_key in each_option.keys():

                                                if option_key == self.XF_label:
                                                    label_dict = each_option[option_key]
                                                    if '#text' in label_dict.keys():
                                                        new_opt['text_label'] = label_dict['#text']
                                                    
                                                    if '@ref' in label_dict.keys():
                                                        new_opt['trans_label'] = label_dict['@ref']

                                                if option_key == self.XF_value:
                                                    value_dict = each_option[option_key]

                                                    if '#text' in value_dict.keys():
                                                        new_opt['text_value'] = value_dict['#text']

                                                    if '@ref' in value_dict.keys():
                                                        new_opt['trans_value'] = value_dict['@ref']

                                            clean_list.append(new_opt)
                                        question['options'] = clean_list
                    



    def writeChoices(self, question, worksheet):

        array_to_write = [None] * 10

        for i in range(0,len(question['options'])):
            cell_to_write = 'A%d' % self.currentRow

            array_to_write[0] = question['element_name']

            if 'text_value' in question['options'][i].keys():
                array_to_write[self.survey_header.index('name')] = question['options'][i]['text_value']
        
            if 'text_label' in question['options'][i].keys():
                
                # print self.survey_header
                # print question['options'][i]['text_label']

                for label_item in self.survey_header:
                    if 'label' in label_item:
                        array_to_write[self.survey_header.index(label_item)] = question['options'][i]['text_label']

                # if 'label' in self.survey_header:
                        


    
            if 'trans_value' in question['options'][i].keys():

                strip_key = question['options'][i]['trans_value'].split("('")[1].split("')")[0]
                trans_obj = self.translation_labels[strip_key]

                'ERROR -> unsupported'

                    # for lang in self.translation_languages:
                        # array_to_write[colheaders.index('hint::%s' % lang)]
                        # print colheaders.index('label::%s' % lang)
                    # array_to_write[self.survey_header('value')] = 


                    # print trans_obj

                    # array_to_write[1] = question['options'][i]['trans_value']
                # else:
                    # array_to_write[1] = '##no entry for value##'

            if 'trans_label' in question['options'][i].keys():
                
                strip_key = question['options'][i]['trans_label'].split("('")[1].split("')")[0]
                # pprint.pprint(self.translation_labels)
                
                # print strip_key

                trans_obj = self.translation_labels[strip_key]
                # print trans_obj
                
                for lang in self.translation_languages:
                    array_to_write[self.survey_header.index('label::%s' % lang)] = trans_obj[lang]

            writeArrayToXLS(worksheet, array_to_write, cell_to_write)

            self.currentRow += 1

        self.currentRow += 1

    def writeXLS(self):

        wb = Workbook()
        worksheet1 = wb.get_active_sheet()
        worksheet1.title = 'survey'

        colheaders = []
        colheaders += ['type', 'name']
        
        if self.number_of_translations > 0:
            for i in range(0,self.number_of_translations):
                # print i
                colheaders += ['label::%s' % self.translation_languages[i]]
                colheaders += ['hint::%s' % self.translation_languages[i]]
                colheaders += ['media::image::%s' % self.translation_languages[i]]
                colheaders += ['media::video::%s' % self.translation_languages[i]]
                colheaders += ['media::audio::%s' % self.translation_languages[i]]
        else:
            colheaders += ['label']
            colheaders += ['hint']
            colheaders += ['media::image']
            colheaders += ['media::video']
            colheaders += ['media::audio']


        colheaders += ['constraint','constraint_message','required','default','relevant','read_only','calculation','appearance']
        writeArrayToXLS(worksheet1, colheaders, 'A1')

        worksheet2 = wb.create_sheet(1)
        worksheet2.title = 'choices'
        self.survey_header = ['list name','name']

        if self.number_of_translations > 0:
            for i in range(0,self.number_of_translations):
                # sprint i
                self.survey_header += ['label::%s' % self.translation_languages[i]]
                # survey_header += ['image::%s' % self.translation_languages[i]]
        else:
            self.survey_header += ['label']
            # survey_header += ['image']

        self.survey_header += ['image']
        
        writeArrayToXLS(worksheet2, self.survey_header, 'A1')

        XF_select1  = "{http://www.w3.org/2002/xforms}select1"
        XF_select   = "{http://www.w3.org/2002/xforms}select"
        XF_input    = "{http://www.w3.org/2002/xforms}input"
        XF_trigger  = "{http://www.w3.org/2002/xforms}trigger"
        XF_group    = "{http://www.w3.org/2002/xforms}group"
        XF_upload   = "{http://www.w3.org/2002/xforms}upload"

        starting_row = 2
        row_offset = 0

        last_group = []

        for question in self.question_list:

            array_to_write = [None] * len(colheaders)

            if 'groups' in question.keys():
                
                group = question['groups']

                if group == last_group:
                    pass


                elif len(group) > len(last_group):
                    # print 'enter group'
                    # moving into subgroup
                    last_group = group
                    array_to_write[colheaders.index('type')] = 'begin group'
                    array_to_write[colheaders.index('name')] = group[0]

                    for glab in self.group_labels:


                        if '@ref' in glab.keys():
                            # array_to_write[colheaders.index('label')] = glab['text_label']
                            print 'unsupported ##'

                        if 'trans_label' in glab.keys():
                            if last_group[-1] in glab['trans_label']:

                                strip_key = glab['trans_label'].split("('")[1].split("')")[0]

                                trans_obj = self.translation_labels[strip_key]
                        
                                for lang in self.translation_languages:
                                    array_to_write[colheaders.index('label::%s' % lang)]
                        
                                    array_to_write[colheaders.index('label::%s' % lang)] = trans_obj[lang]

                        if 'appearance' in glab.keys():
                            print glab.keys()
                    
                    writeArrayToXLS(worksheet1, array_to_write, 'A%d' % int(starting_row+row_offset))
                    row_offset += 1
                    
                else: 
                    # print 'exit gropu'
                    # exiting group
                    last_group = group
                    array_to_write[colheaders.index('type')] = 'end group'
                    
                    writeArrayToXLS(worksheet1, array_to_write, 'A%d' % int(starting_row+row_offset))
                    row_offset += 1

                    # starting a new one
                    if len(group) > 0:

                        array_to_write[colheaders.index('type')] = 'begin group'
                        array_to_write[colheaders.index('name')] = group[0]

                        for glab in self.group_labels:


                            if '@ref' in glab.keys():
                                # array_to_write[colheaders.index('label')] = glab['text_label']
                                print 'unsupported ##'

                            if 'trans_label' in glab.keys():
                                if last_group[-1] in glab['trans_label']:

                                    strip_key = glab['trans_label'].split("('")[1].split("')")[0]

                                    trans_obj = self.translation_labels[strip_key]
                            
                                    for lang in self.translation_languages:
                                        array_to_write[colheaders.index('label::%s' % lang)]
                            
                                        array_to_write[colheaders.index('label::%s' % lang)] = trans_obj[lang]

                            if 'appearance' in glab.keys():
                                print glab.keys()
                        
                        writeArrayToXLS(worksheet1, array_to_write, 'A%d' % int(starting_row+row_offset))
                        row_offset += 1

            elif len(last_group) != 0:

                # print 'exit group'
                    # exiting group
                last_group = []
                array_to_write[colheaders.index('type')] = 'end group'
                
                writeArrayToXLS(worksheet1, array_to_write, 'A%d' % int(starting_row+row_offset))
                row_offset += 1

            for key in question:

                if 'trans_hint' == key:

                    strip_key = question[key].split("('")[1].split("')")[0]
                    trans_obj = self.translation_labels[strip_key]
                    # print trans_obj
                    for lang in self.translation_languages:
                        # array_to_write[colheaders.index('hint::%s' % lang)]
                        # print colheaders.index('hint::%s' % lang)
                        array_to_write[colheaders.index('hint::%s' % lang)] = trans_obj[lang]


                if 'trans_label' == key:

                    strip_key = question[key].split("('")[1].split("')")[0]

                    trans_obj = self.translation_labels[strip_key]
                    # print trans_obj
                    for lang in self.translation_languages:
                        # array_to_write[colheaders.index('hint::%s' % lang)]
                        # print colheaders.index('label::%s' % lang)
                        array_to_write[colheaders.index('label::%s' % lang)] = trans_obj[lang]



                # exact
                if "constraint" == key:
                    array_to_write[colheaders.index('constraint')] = question[key]

                if "element_name" == key:
                    array_to_write[colheaders.index('name')] = question[key]
    
                if "constraint_message" == key:
                    array_to_write[colheaders.index('constraint_message')] = question[key]

                if "default_val" == key:
                    array_to_write[colheaders.index('default')] = question[key]

                if "readonly" == key:
                    array_to_write[colheaders.index('read_only')] = question[key]

                if "calculate" == key:
                    array_to_write[colheaders.index('calculation')] = question[key]

                if 'type' == key:
            
                    type_string = 'text'

                    if question[key] == 'select1':
                        type_string = 'select one from %s' % question['element_name']
                        self.writeChoices(question, worksheet2)

                    elif question[key] == 'select':
                        type_string = 'select_multiple %s' % question['element_name']
                        self.writeChoices(question, worksheet2)

                    elif question[key] == 'binary':
                        type_string = 'photo'

                    elif question[key] == 'undefined':
                        type_string = 'text'

                    elif question[key] == 'int':
                        type_string = 'integer'
                    else:
                        type_string = question[key]

                    array_to_write[colheaders.index('type')] = type_string

                if 'required' == key:
                    array_to_write[colheaders.index('required')] = question[key]

            # pprint.pprint(question)
            # print '\n'

            writeArrayToXLS(worksheet1, array_to_write, 'A%d' % int(starting_row+row_offset))
            row_offset += 1

        ## in case we end on a group
        if len(last_group) != 0:

            last_group = []
            array_to_write = [None] * len(colheaders)
            array_to_write[colheaders.index('type')] = 'end group'
            writeArrayToXLS(worksheet1, array_to_write, 'A%d' % int(starting_row+row_offset))



        worksheet3 = wb.create_sheet(2)
        worksheet3.title = 'settings'
        settings_header = ['form_title', 'form_id', 'public_key', 'submission_url', 'default_language']
        writeArrayToXLS(worksheet3, settings_header, 'A1')

        settings_array_to_write = [self.title, '', '', '', self.default_lang ]
        writeArrayToXLS(worksheet3, settings_array_to_write, 'A2')

        wb.save(self.output_file)



if __name__ == '__main__':

    # file_list = ['census.json', 'risk.json', 'satis.json', 'trauma.json']
    file_list = ['nhrcnew.xml']

    for file_to_use in file_list:
        print 'converting %s' % file_to_use
        converter = xformConvert(file_to_use)
        converter.parseTranslations()
        converter.parseBindOrder()
        converter.parseInstanceData()
        converter.parseBody()
        converter.writeXLS()


